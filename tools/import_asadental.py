import os
import sys
import json
import re
import openpyxl

# 1. Install Pillow if not present
try:
    from PIL import Image
    print("Pillow is installed")
except ImportError:
    print("Pillow is not installed, trying to install...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "Pillow"])
    from PIL import Image

# Config paths
XLSX_PATH = r"C:\Users\garbarking\Downloads\Pricing_Megastandard_TR_IQ.xlsx"
CAT1_DIR  = r"C:\Users\garbarking\Downloads\wetransfer_half1_catalogue_2026-07-03_1341\Half1_Catalogue"
CAT2_DIR  = r"C:\Users\garbarking\Downloads\wetransfer_half2_catalogue_2026-07-03_1346\Half2_Catalogue"

PROJECT_ROOT = r"C:\Users\garbarking\.gemini\antigravity\scratch\tammuz-dental"
OUTPUT_IMG_DIR = os.path.join(PROJECT_ROOT, "assets", "images", "products", "asadental")
DATA_JSON_PATH = os.path.join(PROJECT_ROOT, "products", "data.json")

# Category mapping
FOLDER_TO_CAT = {
    'A_AsaOne disposables':                  ('auxiliary', 'Preventive & Auxiliary'),
    'B_Diagnostic':                          ('diagnostic', 'Diagnostics'),
    'C_Oral Surgery':                        ('surgery', 'Oral & Implant Surgery'),
    'D_Extractive Surgery':                  ('surgery', 'Oral & Implant Surgery'),
    'E_Implant Surgery':                     ('surgery', 'Oral & Implant Surgery'),
    'Ideal Periotomi':                       ('surgery', 'Oral & Implant Surgery'),
    'F_Restorative':                         ('restorative', 'Restoratives'),
    'G_Periodontal':                         ('periodontal', 'Periodontics'),
    'H_Orthodontic':                         ('orthodontic', 'Orthodontics'),
    'I_Instrument cassettes and trays':      ('trays', 'Cassettes & Trays'),
    'J_Impression Trays':                    ('impression', 'Impression'),
    'K_Laboratory instruments':              ('laboratory', 'Laboratory Instruments')
}

def normalize(text):
    if not text:
        return ""
    t = text.lower()
    t = re.sub(r'[\s\-_/\\#]', '', t)
    return t

print("Scanning image files...")
image_files = []
for cat_dir in [CAT1_DIR, CAT2_DIR]:
    if not os.path.exists(cat_dir):
        print(f"Directory not found: {cat_dir}")
        continue
    for root, dirs, files in os.walk(cat_dir):
        for f in files:
            ext = os.path.splitext(f)[1].lower()
            if ext in {'.jpeg', '.jpg', '.png', '.webp'}:
                full_path = os.path.join(root, f)
                folder_name = os.path.basename(os.path.dirname(full_path))
                image_files.append({
                    'filename': f,
                    'name_no_ext': os.path.splitext(f)[0].strip(),
                    'full_path': full_path,
                    'folder_name': folder_name
                })

print(f"Total image files found: {len(image_files)}")

# Map normalized image names to their metadata
image_by_norm = {}
for img in image_files:
    norm_name = normalize(img['name_no_ext'])
    image_by_norm[norm_name] = img
    
    # Split names like "3251E-B-9 3252E-B-9"
    parts = re.split(r'[\s,]+', img['name_no_ext'])
    for p in parts:
        norm_p = normalize(p)
        if norm_p and norm_p not in image_by_norm:
            image_by_norm[norm_p] = img

print("Loading pricing Excel...")
wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
excel_items = {}

# We'll merge items from all sheets. The last one will overwrite or we can average,
# but they are probably identical since they represent lists of prices.
# 'DEMOZI TURKIA' is the active distributor sheet for Turkey.
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        continue
    header = rows[0]
    try:
        idx_code = header.index('Item Code')
        idx_desc = header.index('Description')
        idx_price = header.index('Net Price EX WORKS / EURO')
    except ValueError:
        print(f"Skipping sheet {sheet_name} as headers mismatch: {header}")
        continue
        
    for r in rows[1:]:
        code = str(r[idx_code]).strip() if r[idx_code] is not None else None
        desc = str(r[idx_desc]).strip() if r[idx_desc] is not None else ""
        price = r[idx_price]
        if not code:
            continue
            
        # Standardize price formatting
        price_val = None
        if price is not None:
            try:
                price_val = float(price)
            except ValueError:
                pass
                
        if code not in excel_items:
            excel_items[code] = {
                'code': code,
                'description': desc,
                'price': price_val
            }

print(f"Unique item codes in Excel: {len(excel_items)}")

# Match items
matched_items = []
for code, item in excel_items.items():
    norm_code = normalize(code)
    if norm_code in image_by_norm:
        img_info = image_by_norm[norm_code]
        item['image_src_path'] = img_info['full_path']
        item['folder_name'] = img_info['folder_name']
        matched_items.append(item)

print(f"Matched products: {len(matched_items)} / {len(excel_items)}")

# Create output image folder
os.makedirs(OUTPUT_IMG_DIR, exist_ok=True)

# We'll process each matched item, resize/compress its image, and generate JSON.
new_products_json = []
processed_count = 0

print("Processing images and building catalog entries (resizing to max 800px and saving as WebP)...")
for idx, item in enumerate(matched_items):
    code = item['code']
    description = item['description']
    price = item['price']
    src_img_path = item['image_src_path']
    folder = item['folder_name']
    
    # Map category
    cat_id, cat_label = FOLDER_TO_CAT.get(folder, ('auxiliary', 'Preventive & Auxiliary'))
    
    # Generate clean ID and filename
    code_slug = re.sub(r'[^a-zA-Z0-9-]', '_', code)
    dest_filename = f"asa_{code_slug}.webp"
    dest_img_path = os.path.join(OUTPUT_IMG_DIR, dest_filename)
    rel_img_path = f"assets/images/products/asadental/{dest_filename}"
    
    # Resize and save image
    try:
        if not os.path.exists(dest_img_path):
            with Image.open(src_img_path) as img:
                # Convert to RGB if it is RGBA/P
                if img.mode in ("RGBA", "P"):
                    img = img.convert("RGB")
                # Thumbnail maintains aspect ratio
                img.thumbnail((800, 800))
                img.save(dest_img_path, "WEBP", quality=80)
    except Exception as e:
        print(f"Error processing image {src_img_path}: {e}")
        # fallback to original extension copy or ignore
        continue

    # Create price display text
    price_text = f"€{price:.2f}" if price is not None else "Request Quote"

    # Assemble specs
    specs = [
        f"Item Code: {code}",
        f"Manufacturer: Asa Dental (Italy)",
        f"Category: {cat_label}"
    ]
    if price is not None:
        specs.append(f"Distributor B2B Price: {price_text} EXW")

    product_entry = {
        "id": f"asa-{code_slug.lower()}",
        "category": cat_id,
        "category_label": cat_label,
        "name": description or f"Asa Dental Instrument {code}",
        "name_tr": description or f"Asa Dental Enstrümanı {code}",
        "description": f"Premium Italian-made dental instrument by Asa Dental. Certified quality for professional practice.",
        "description_tr": "Asa Dental tarafından İtalya'da üretilmiş birinci sınıf dental enstrüman. Profesyonel uygulamalar için sertifikalı kalite.",
        "unit": "1 Unit",
        "image": rel_img_path,
        "specs": specs,
        "tags": ["asadental", cat_id, code.lower()]
    }
    
    new_products_json.append(product_entry)
    processed_count += 1
    if processed_count % 200 == 0:
        print(f"  Processed {processed_count}/{len(matched_items)} items...")

# Load existing products
existing_products = []
if os.path.exists(DATA_JSON_PATH):
    try:
        with open(DATA_JSON_PATH, 'r', encoding='utf-8') as f:
            existing_products = json.load(f)
        print(f"Loaded {len(existing_products)} existing products from data.json")
    except Exception as e:
        print(f"Error reading existing data.json: {e}")

# Avoid duplicates if matching Asadental IDs are already present
existing_ids = {p['id'] for p in existing_products}
filtered_new_products = [p for p in new_products_json if p['id'] not in existing_ids]

# Merge
final_products = existing_products + filtered_new_products
print(f"Total products in new catalog: {len(final_products)} (Added {len(filtered_new_products)} new Asadental items)")

# Save back to data.json
with open(DATA_JSON_PATH, 'w', encoding='utf-8') as f:
    json.dump(final_products, f, indent=2, ensure_ascii=False)

print("Catalog JSON file successfully updated!")
